import os
from openpyxl import load_workbook

dir_search = "./Archivos_vinculados"
string_match = "Fila"
string_replace = "Nueva Fila"
extensions = [".xls",".xlsx"]
files_procesed = 0
files_matched = 0

for root, dirs, files in os.walk(dir_search, True):
  if files:
    files_matched = 0
    for file in files:
      filename, file_extension = os.path.splitext(file)
      if file_extension in extensions:
        files_procesed+=1
        book = load_workbook(os.path.join(root,file))
        for sheet in book:
          for row in sheet:
            for cell in row:
              ##print(cell.value)
              if cell.value is not None:
                valor_cell = cell.value
                if string_match in cell.value:
                  files_matched+=1
                  ##cell.value = valor_cell.replace(string_match, string_replace)
              if cell.hyperlink is not None:
                print(f"Tengo valor hyperlink {cell.hyperlink}")
          # print(f"Book -> {book.sheetnames}: ")
          # print(f"Path -> {root}: ")
          # print(f"File: {file} con extension: {file_extension}")
        print(f"Se encontraron {files_matched} coincidencias en archivo {filename}")
        book.save(filename=os.path.join(root,file))
print(f"{files_procesed} arhivos con extension {str(extensions)} procesados")